# -*- coding: utf-8 -*-
"""
Generate a monthly literature statistics workbook from classified weekly Excel files.

Default behavior:
  - Month: previous month in Asia/Shanghai, e.g. 2026-06-03 -> 2026-05.
  - Input: output/weekly/*_translated.xlsx
  - Output: output/monthly/YYYY-MM 文献统计表.xlsx
"""

from __future__ import annotations

import argparse
import calendar
import re
import shutil
from copy import copy
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Iterable
from zoneinfo import ZoneInfo

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_WEEKLY_DIR = BASE_DIR / "output" / "weekly"
DEFAULT_OUTPUT_DIR = BASE_DIR / "output" / "monthly"
DEFAULT_SOURCE_MAP = BASE_DIR / "config" / "monthly" / "source对应表.xlsx"
DEFAULT_TEMPLATE = BASE_DIR / "config" / "monthly" / "文献统计模板表.xlsx"

NATURE_COMMUNICATIONS_SOURCES = {
    "Nature Communications",
    "Biological sciences : Nature Communications subject feeds",
    "Scientific community and society : Nature Communications subject feeds",
    "Earth and environmental sciences : Nature Communications subject feeds",
    "Physical sciences : Nature Communications subject feeds",
}
NATURE_COMMUNICATIONS_CANONICAL_SOURCE = "Nature Communications"
ASAP_PREFIX = "[ASAP] "

SHEET_ALIASES = {
    "氢基能源": ["氢基能源", "氢氨醇"],
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate monthly literature statistics workbook.")
    parser.add_argument(
        "--month",
        help="Target month in YYYY-MM. Defaults to the previous month in Asia/Shanghai.",
    )
    parser.add_argument("--weekly-dir", type=Path, default=DEFAULT_WEEKLY_DIR)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--source-map", type=Path, default=DEFAULT_SOURCE_MAP)
    parser.add_argument("--template", type=Path, default=DEFAULT_TEMPLATE)
    parser.add_argument(
        "--include-untranslated",
        action="store_true",
        help="Also read weekly .xlsx files that do not end with _translated.xlsx.",
    )
    return parser.parse_args()


def default_previous_month() -> str:
    now = datetime.now(ZoneInfo("Asia/Shanghai"))
    first_of_this_month = now.replace(day=1)
    last_of_previous_month = first_of_this_month - timedelta(days=1)
    return last_of_previous_month.strftime("%Y-%m")


def month_bounds(month: str) -> tuple[pd.Timestamp, pd.Timestamp]:
    if not re.fullmatch(r"\d{4}-\d{2}", month):
        raise ValueError("--month must use YYYY-MM format")
    year, mon = map(int, month.split("-"))
    last_day = calendar.monthrange(year, mon)[1]
    start = pd.Timestamp(f"{month}-01", tz="UTC")
    end = pd.Timestamp(datetime(year, mon, last_day, 23, 59, 59, tzinfo=timezone.utc))
    return start, end


def normalize_source(source) -> str:
    if pd.isna(source):
        return ""
    value = str(source).strip()
    if value in NATURE_COMMUNICATIONS_SOURCES:
        return NATURE_COMMUNICATIONS_CANONICAL_SOURCE
    return value


def clean_title(title) -> str:
    if pd.isna(title):
        return ""
    value = str(title).strip()
    if value.startswith(ASAP_PREFIX):
        return value[len(ASAP_PREFIX) :].strip()
    return value


def first_nonempty(*values) -> str:
    for value in values:
        if pd.isna(value):
            continue
        text = str(value).strip()
        if text:
            return text
    return ""


def display_len(value) -> int:
    if value is None:
        return 0
    return sum(2 if ord(ch) > 127 else 1 for ch in str(value))


def set_left_wrap_alignment(cell) -> None:
    cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)


def copy_cell_style(source_cell, target_cell) -> None:
    if not source_cell.has_style:
        return
    target_cell.font = copy(source_cell.font)
    target_cell.border = copy(source_cell.border)
    target_cell.fill = copy(source_cell.fill)
    target_cell.number_format = copy(source_cell.number_format)
    target_cell.alignment = copy(source_cell.alignment)


def adjust_sheet_layout(ws, headers: dict[str, int]) -> None:
    wide_headers = {"标题": 62, "DOI": 52}
    preferred_widths = {
        "出版商": 16,
        "期刊名": 26,
        "通讯作者": 18,
        "发表日期": 14,
        "数量": 8,
    }

    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col_idx).value
        header = str(header).strip() if header else ""
        letter = openpyxl.utils.get_column_letter(col_idx)

        if header in wide_headers:
            ws.column_dimensions[letter].width = wide_headers[header]
        elif header in preferred_widths:
            ws.column_dimensions[letter].width = preferred_widths[header]
        else:
            max_len = display_len(header)
            for row_idx in range(2, ws.max_row + 1):
                max_len = max(max_len, display_len(ws.cell(row=row_idx, column=col_idx).value))
            ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 22)

    for row in ws.iter_rows():
        for cell in row:
            set_left_wrap_alignment(cell)

    for row_idx in range(1, ws.max_row + 1):
        if row_idx == 1:
            ws.row_dimensions[row_idx].height = 24
            continue

        max_lines = 1
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is None:
                continue
            letter = openpyxl.utils.get_column_letter(col_idx)
            width = ws.column_dimensions[letter].width or 12
            max_lines = max(max_lines, int(display_len(cell.value) / max(width, 1)) + 1)

        ws.row_dimensions[row_idx].height = min(max(18, max_lines * 16), 96)


def load_source_map(path: Path) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(f"Source map not found: {path}")

    df = pd.read_excel(path, keep_default_na=False)
    required = {"source", "出版社", "期刊名"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"Source map missing columns: {', '.join(sorted(missing))}")

    df = df[["source", "出版社", "期刊名"]].copy()
    df["source"] = df["source"].apply(normalize_source)
    df["出版社"] = df["出版社"].astype(str).str.strip()
    df["期刊名"] = df["期刊名"].astype(str).str.strip()
    return df.drop_duplicates(subset=["source"], keep="first")


def weekly_files_for_month(weekly_dir: Path, month: str, include_untranslated: bool) -> list[Path]:
    if not weekly_dir.exists():
        raise FileNotFoundError(f"Weekly directory not found: {weekly_dir}")

    pattern = f"weekly_news_with_abstract_{month}-*.xlsx"
    files = sorted(weekly_dir.glob(pattern))
    if not include_untranslated:
        files = [path for path in files if path.name.endswith("_translated.xlsx")]

    return [path for path in files if "classify_only" not in path.name]


def read_weekly_sheet(path: Path, sheet_name: str) -> pd.DataFrame:
    xls = pd.ExcelFile(path)
    aliases = SHEET_ALIASES.get(sheet_name, [sheet_name])
    target_sheet = next((name for name in aliases if name in xls.sheet_names), None)
    if not target_sheet:
        return pd.DataFrame()

    df = pd.read_excel(xls, sheet_name=target_sheet, keep_default_na=False)
    if df.empty:
        return df

    df["__weekly_file"] = path.name
    return df


def normalize_article_data(df: pd.DataFrame, source_map: pd.DataFrame, start: pd.Timestamp, end: pd.Timestamp) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()

    df = df.copy()
    for column in ["title", "title_zh", "source", "pub_date", "published", "doi", "link", "last_author"]:
        if column not in df.columns:
            df[column] = ""

    df["source"] = df["source"].apply(normalize_source)
    df["pub_date"] = pd.to_datetime(df["pub_date"], utc=True, errors="coerce")
    missing_pub_date = df["pub_date"].isna()
    if missing_pub_date.any():
        df.loc[missing_pub_date, "pub_date"] = pd.to_datetime(
            df.loc[missing_pub_date, "published"],
            utc=True,
            errors="coerce",
        )

    df = df[df["pub_date"].notna() & (df["pub_date"] >= start) & (df["pub_date"] <= end)].copy()
    if df.empty:
        return df

    merged = pd.merge(df, source_map, on="source", how="inner")
    if merged.empty:
        return merged

    merged["__stable_key"] = merged.apply(
        lambda row: first_nonempty(row.get("stable_id", ""), row.get("link", ""), row.get("doi", ""), row.get("title", "")),
        axis=1,
    )
    merged = merged.drop_duplicates(subset=["__stable_key", "出版社", "期刊名"], keep="first")
    merged = merged.sort_values(["出版社", "期刊名", "pub_date", "title"], kind="stable")
    return merged


def headers_for(ws) -> dict[str, int]:
    return {str(cell.value).strip(): idx for idx, cell in enumerate(ws[1], start=1) if cell.value}


def unmerge_template_columns(ws, columns: Iterable[int]) -> None:
    target_columns = set(col for col in columns if col)
    for merged_range in list(ws.merged_cells.ranges):
        if (
            merged_range.min_col == merged_range.max_col
            and merged_range.min_col in target_columns
        ):
            ws.unmerge_cells(str(merged_range))


def fill_down_publisher_cells(ws, col_pub: int) -> None:
    current_value = None
    current_style_cell = None
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=col_pub)
        if cell.value:
            current_value = str(cell.value).strip()
            current_style_cell = cell
        else:
            cell.value = current_value
            if current_style_cell:
                copy_cell_style(current_style_cell, cell)


def template_rows(ws, col_pub: int, col_jour: int) -> list[tuple[int, str, str]]:
    rows = []
    for row_idx in range(2, ws.max_row + 1):
        pub_value = ws.cell(row=row_idx, column=col_pub).value
        jour_value = ws.cell(row=row_idx, column=col_jour).value
        pub = str(pub_value).strip() if pub_value else ""
        jour = str(jour_value).strip() if jour_value else ""
        if jour:
            rows.append((row_idx, pub, jour))
    return rows


def merge_equal_runs(ws, col_idx: int, key_columns: list[int]) -> None:
    start_row = 2
    current_key = tuple(ws.cell(row=2, column=col).value for col in key_columns)

    for row_idx in range(3, ws.max_row + 2):
        if row_idx <= ws.max_row:
            row_key = tuple(ws.cell(row=row_idx, column=col).value for col in key_columns)
        else:
            row_key = (None,) * len(key_columns)

        if row_key != current_key:
            if row_idx - 1 > start_row and ws.cell(row=start_row, column=col_idx).value:
                ws.merge_cells(
                    start_row=start_row,
                    start_column=col_idx,
                    end_row=row_idx - 1,
                    end_column=col_idx,
                )
                set_left_wrap_alignment(ws.cell(row=start_row, column=col_idx))
            start_row = row_idx
            current_key = row_key


def fill_sheet(ws, articles: pd.DataFrame) -> None:
    headers = headers_for(ws)
    col_pub = headers.get("出版商")
    col_jour = headers.get("期刊名")
    col_title = headers.get("标题")
    col_author = headers.get("通讯作者")
    col_date = headers.get("发表日期")
    col_doi = headers.get("DOI")
    col_count = headers.get("数量")

    if not col_pub or not col_jour:
        return

    unmerge_template_columns(ws, [col_pub, col_jour, col_count])
    fill_down_publisher_cells(ws, col_pub)
    rows = template_rows(ws, col_pub, col_jour)

    for row_idx, publisher, journal in reversed(rows):
        if not articles.empty:
            matched = articles[(articles["出版社"] == publisher) & (articles["期刊名"] == journal)]
        else:
            matched = pd.DataFrame()

        count = len(matched)
        if count == 0:
            if col_count:
                ws.cell(row=row_idx, column=col_count).value = 0
            continue

        if count > 1:
            ws.insert_rows(row_idx + 1, count - 1)
            for new_row_idx in range(row_idx + 1, row_idx + count):
                for col_idx in range(1, ws.max_column + 1):
                    copy_cell_style(ws.cell(row=row_idx, column=col_idx), ws.cell(row=new_row_idx, column=col_idx))

        for offset, (_, article) in enumerate(matched.iterrows()):
            current_row = row_idx + offset
            pub_date = article.get("pub_date", "")
            if pd.notna(pub_date):
                pub_date = pd.Timestamp(pub_date).strftime("%Y-%m-%d")
            else:
                pub_date = ""

            title = first_nonempty(article.get("title", ""), article.get("title_zh", ""))
            doi_or_link = first_nonempty(article.get("doi", ""), article.get("link", ""))

            ws.cell(row=current_row, column=col_pub).value = publisher
            ws.cell(row=current_row, column=col_jour).value = journal
            if col_title:
                ws.cell(row=current_row, column=col_title).value = clean_title(title)
            if col_author:
                ws.cell(row=current_row, column=col_author).value = first_nonempty(article.get("last_author", ""))
            if col_date:
                ws.cell(row=current_row, column=col_date).value = pub_date
            if col_doi:
                ws.cell(row=current_row, column=col_doi).value = doi_or_link
            if col_count:
                ws.cell(row=current_row, column=col_count).value = count

    merge_equal_runs(ws, col_jour, [col_pub, col_jour])
    if col_count:
        merge_equal_runs(ws, col_count, [col_pub, col_jour])
    merge_equal_runs(ws, col_pub, [col_pub])
    adjust_sheet_layout(ws, headers)


def generate_monthly_workbook(
    month: str,
    weekly_dir: Path,
    output_dir: Path,
    source_map_path: Path,
    template_path: Path,
    include_untranslated: bool = False,
) -> Path:
    start, end = month_bounds(month)
    source_map = load_source_map(source_map_path)

    if not template_path.exists():
        raise FileNotFoundError(f"Template workbook not found: {template_path}")

    weekly_files = weekly_files_for_month(weekly_dir, month, include_untranslated)
    if not weekly_files:
        raise FileNotFoundError(f"No weekly Excel files found for {month} in {weekly_dir}")

    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"{month} 文献统计表.xlsx"
    shutil.copy(template_path, output_path)

    wb = openpyxl.load_workbook(output_path)
    for sheet_name in wb.sheetnames:
        sheet_frames = []
        for weekly_file in weekly_files:
            frame = read_weekly_sheet(weekly_file, sheet_name)
            if not frame.empty:
                sheet_frames.append(frame)

        raw_articles = pd.concat(sheet_frames, ignore_index=True) if sheet_frames else pd.DataFrame()
        articles = normalize_article_data(raw_articles, source_map, start, end)
        fill_sheet(wb[sheet_name], articles)
        print(f"[monthly] {sheet_name}: {len(articles)} matched records")

    wb.save(output_path)
    print(f"[monthly] Read {len(weekly_files)} weekly workbook(s)")
    print(f"[monthly] Wrote {output_path}")
    return output_path


def main() -> None:
    args = parse_args()
    month = args.month or default_previous_month()
    generate_monthly_workbook(
        month=month,
        weekly_dir=args.weekly_dir,
        output_dir=args.output_dir,
        source_map_path=args.source_map,
        template_path=args.template,
        include_untranslated=args.include_untranslated,
    )


if __name__ == "__main__":
    main()
